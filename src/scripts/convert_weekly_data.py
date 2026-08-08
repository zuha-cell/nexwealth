"""
Converts the weekly "MF & ETFs Ready Reckoner" workbook into clean JSON,
one record per fund, ready to import into the `mutualFunds` Firestore collection.

Usage:
    python convert_weekly_data.py "MF & ETFs Weekly Ready Reckoner as on <date>.xlsx" output.json

The workbook has one sheet per fund category (Large Cap, Mid Cap, ETF, ...).
Each sheet follows the same layout:
    row 3 (0-indexed 2): report title
    row 4 (0-indexed 3): category label
    row 5 (0-indexed 4): top-level column groups (Scheme Name, ISIN, NAV, ...,
                          "Simple Absolute", "CAGR", "SIP Returns", ...)
    row 6 (0-indexed 5): sub-labels under those groups (1M, 3M, 1 Yr, 3 Yr, ...)
    row 7+ (0-indexed 6+): fund data

Sheets that aren't per-category fund tables (Home, Glossary, Disclaimer, etc.)
are skipped automatically.
"""

import json
import sys
from datetime import datetime
import openpyxl

# Sheets that are NOT category fund tables — skip these.
SKIP_SHEETS = {
    "Template ID", "Home", "Mirae Asset Schemes", "Mutual Fund Top Picks",
    "SIF", "Category Performance",
    "Benchmark", "Glossary", "Disclaimer", "Sheet2",
}

# Known "core" columns that appear before the returns-metric columns start,
# in the order they appear in row 5. Anything after these is a returns metric.
CORE_FIELDS = [
    "fundCode", "schemeName", "isin", "nav", "schemeType",
    "inceptionDate", "ageYears", "aumCr", "expenseRatio",
]


def clean(v):
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    return v


def parse_sheet(ws, category):
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 7:
        return []

    group_row = rows[4]   # e.g. "Simple Absolute", "CAGR", "SIP Returns", "Asset Allocation"...
    label_row = rows[5]   # e.g. "1M", "3M", "1 Yr", ...

    # Build column -> metric key map for everything after the core columns.
    # Forward-fill the group label across merged cells.
    last_group = None
    col_meta = {}
    for col_idx in range(len(group_row)):
        g = group_row[col_idx]
        if g is not None:
            last_group = str(g).strip()
        label = label_row[col_idx]
        if label is not None and last_group:
            key = f"{last_group}:{str(label).strip()}"
            col_meta[col_idx] = key

    funds = []
    for row in rows[6:]:
        if row is None or row[1] is None:  # no scheme name -> not a data row
            continue
        record = {
            "category": category,
            "fundCode": clean(row[0]) if len(row) > 0 else None,
            "schemeName": clean(row[1]) if len(row) > 1 else None,
            "isin": clean(row[2]) if len(row) > 2 else None,
            "nav": clean(row[3]) if len(row) > 3 else None,
            "schemeType": clean(row[4]) if len(row) > 4 else None,
            "inceptionDate": clean(row[5]) if len(row) > 5 else None,
            "ageYears": clean(row[6]) if len(row) > 6 else None,
            "aumCr": clean(row[7]) if len(row) > 7 else None,
            "expenseRatio": clean(row[8]) if len(row) > 8 else None,
            "metrics": {},
        }
        for col_idx, key in col_meta.items():
            if col_idx < 9:
                continue  # already captured as a core field
            if col_idx < len(row):
                val = clean(row[col_idx])
                if val is not None:
                    record["metrics"][key] = val
        funds.append(record)
    return funds


def main():
    if len(sys.argv) != 3:
        print("Usage: python convert_weekly_data.py <input.xlsx> <output.json>")
        sys.exit(1)

    src, dest = sys.argv[1], sys.argv[2]
    wb = openpyxl.load_workbook(src, read_only=True, data_only=True)

    all_funds = []
    summary = {}
    for name in wb.sheetnames:
        if name in SKIP_SHEETS:
            continue
        ws = wb[name]
        funds = parse_sheet(ws, category=name.strip())
        if funds:
            all_funds.extend(funds)
            summary[name.strip()] = len(funds)

    with open(dest, "w") as f:
        json.dump(all_funds, f, indent=2, default=str)

    print(f"Wrote {len(all_funds)} fund records across {len(summary)} categories to {dest}")
    for cat, count in summary.items():
        print(f"  {cat}: {count}")


if __name__ == "__main__":
    main()
